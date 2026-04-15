import pandas as pd, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
R = lambda p: pd.read_csv(os.path.join(ROOT, p), encoding="utf-8-sig")

att = R("attendance/kbo_attendance_1982_2025.csv")
ks = R("postseason/korean_series_1982_2025.csv")
mvp = R("awards/kbo_mvp_1982_2025.csv")
gg = R("awards/kbo_golden_glove_1982_2025.csv")
st = R("team/team_standings_2001_2025.csv")

wb = Workbook()
wb.remove(wb.active)

NAVY = "1F2937"; GRAY = "F3F4F6"
TITLE = Font(name="맑은 고딕", size=16, bold=True, color="111827")
SUB   = Font(name="맑은 고딕", size=10, color="6B7280")
HF    = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BF    = Font(name="맑은 고딕", size=10, color="111827")
HFILL = PatternFill("solid", fgColor=NAVY)
THIN  = Side(style="thin", color="D1D5DB")
BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN   = Alignment(horizontal="center", vertical="center")
LFT   = Alignment(horizontal="left", vertical="center")
RGT   = Alignment(horizontal="right", vertical="center")


def header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HF; c.fill = HFILL; c.alignment = CEN; c.border = BORD
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def body(ws, row, values, aligns):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BF; c.border = BORD; c.alignment = aligns[i - 1]


# Sheet 1: 요약
s1 = wb.create_sheet("요약"); s1.sheet_view.showGridLines = False
s1.column_dimensions["A"].width = 2
s1["B2"] = "KBO 2025 시즌 리뷰"; s1["B2"].font = TITLE
s1["B3"] = "작성: 기획팀  |  데이터 기준: 2025 정규시즌 종료  |  출처: KBO 공식·Wikipedia"
s1["B3"].font = SUB
s1.merge_cells("B2:G2"); s1.merge_cells("B3:G3")

y25 = att[att.year == 2025].iloc[0]
y24 = att[att.year == 2024].iloc[0]
y19 = att[att.year == 2019].iloc[0]
kpis = [
    ("총 관중", f"{int(y25.total_attendance):,}명",
     f"전년 대비 +{(y25.total_attendance / y24.total_attendance - 1) * 100:.1f}%"),
    ("경기당 평균", f"{int(y25.avg_per_game):,}명",
     f"COVID 前(2019) 대비 +{(y25.avg_per_game / y19.avg_per_game - 1) * 100:.0f}%"),
    ("우승팀", ks[ks.year == 2025].champion.iloc[0],
     f"한국시리즈 {ks[ks.year == 2025].series_result.iloc[0]}"),
    ("정규시즌 MVP", mvp[mvp.year == 2025].player.iloc[0],
     f"{mvp[mvp.year == 2025].team.iloc[0]}, {mvp[mvp.year == 2025].position.iloc[0]}"),
]
CFILL = PatternFill("solid", fgColor=GRAY)
for i, (lab, val, note) in enumerate(kpis):
    col = 2 + (i % 2) * 3
    row = 5 + (i // 2) * 4
    s1.cell(row=row, column=col, value=lab).font = Font(name="맑은 고딕", size=9, color="6B7280")
    s1.cell(row=row + 1, column=col, value=val).font = Font(name="맑은 고딕", size=14, bold=True, color="111827")
    s1.cell(row=row + 2, column=col, value=note).font = SUB
    for rr in range(row, row + 3):
        for cc in range(col, col + 3):
            s1.cell(row=rr, column=cc).fill = CFILL
    for rr in range(row, row + 3):
        s1.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=col + 2)

for letter in ["B", "C", "D", "E", "F", "G"]:
    s1.column_dimensions[letter].width = 18

s1["B14"] = "핵심 관찰"; s1["B14"].font = Font(name="맑은 고딕", size=12, bold=True)
bullets = [
    "• COVID(2020) 32.8만 관중 → 2025 1,231만 관중, 역대 최초 1,200만 돌파",
    "• LG Twins 2023 이후 2년 연속 Top 3, 2025 정규시즌 1위",
    "• 한화 Eagles 2021 10위 → 2025 2위, 5년간 변동폭 최대",
    "• 외국인 투수의 정규시즌 MVP 수 년 연속 — 선발 이닝 외인 의존 구조 지속",
]
for i, t in enumerate(bullets):
    c = s1.cell(row=15 + i, column=2, value=t)
    c.font = Font(name="맑은 고딕", size=10, color="374151")
    s1.merge_cells(start_row=15 + i, start_column=2, end_row=15 + i, end_column=7)

# Sheet 2: 2025 순위
s2 = wb.create_sheet("2025순위"); s2.sheet_view.showGridLines = False
s2["A1"] = "2025 정규시즌 최종 순위"; s2["A1"].font = TITLE
s2.merge_cells("A1:H1")
header(s2, 3, ["순위", "팀", "경기", "승", "패", "무", "승률", "게임차"], [8, 22, 8, 8, 8, 8, 10, 10])
s25 = st[st.year == 2025].sort_values("rank")
for i, r in enumerate(s25.itertuples(index=False), start=4):
    body(s2, i, [r.rank, r.team, int(r.G), int(r.W), int(r.L), int(r.D), r.WPCT, r.GB],
         [CEN, LFT, RGT, RGT, RGT, RGT, RGT, RGT])
    s2.cell(row=i, column=7).number_format = "0.000"
s2.conditional_formatting.add(f"G4:G{3 + len(s25)}",
    DataBarRule(start_type="min", end_type="max", color="3B82F6"))

# Sheet 3: 관중 추세
s3 = wb.create_sheet("관중추세"); s3.sheet_view.showGridLines = False
s3["A1"] = "KBO 총 관중 1982-2025"; s3["A1"].font = TITLE
s3.merge_cells("A1:F1")
header(s3, 3, ["연도", "총관중", "경기당평균", "경기수"], [10, 14, 14, 10])
for i, r in enumerate(att.itertuples(index=False), start=4):
    body(s3, i, [r.year, r.total_attendance, r.avg_per_game, r.number_of_games],
         [CEN, RGT, RGT, RGT])
    for cc in [2, 3, 4]:
        s3.cell(row=i, column=cc).number_format = "#,##0"
ch = LineChart(); ch.title = "연도별 총 관중"; ch.height = 9; ch.width = 18
data_ref = Reference(s3, min_col=2, min_row=3, max_row=3 + len(att))
cat_ref = Reference(s3, min_col=1, min_row=4, max_row=3 + len(att))
ch.add_data(data_ref, titles_from_data=True); ch.set_categories(cat_ref)
s3.add_chart(ch, "F3")

# Sheet 4: 골든글러브 최다
s4 = wb.create_sheet("GG_최다수상"); s4.sheet_view.showGridLines = False
s4["A1"] = "골든글러브 최다 수상 선수 (1982-2025)"; s4["A1"].font = TITLE
s4.merge_cells("A1:D1")
top = gg.player.value_counts().head(15).reset_index()
top.columns = ["player", "count"]
header(s4, 3, ["순위", "선수", "수상횟수"], [8, 22, 12])
for i, r in enumerate(top.itertuples(index=False), start=4):
    body(s4, i, [i - 3, r.player, r.count], [CEN, LFT, RGT])
s4.conditional_formatting.add(f"C4:C{3 + len(top)}",
    ColorScaleRule(start_type="min", start_color="DBEAFE", end_type="max", end_color="1E40AF"))

# Sheet 5: 데이터 출처
s5 = wb.create_sheet("데이터_출처"); s5.sheet_view.showGridLines = False
s5["A1"] = "데이터 출처 및 한계"; s5["A1"].font = TITLE
s5.merge_cells("A1:B1")
rows = [
    ("카테고리", "출처"),
    ("선수·팀 스탯 (2001-2020)", "GitHub baseballChatbot7/KBO-Record-Crawler (KBO 공식 기록실 크롤링)"),
    ("팀 순위 (2021-2025)", "Wikipedia — 각 시즌 KBO League season 페이지"),
    ("관중 (1982-2025)", "Wikipedia — KBO League 본문"),
    ("한국시리즈 (1982-2025)", "Wikipedia — Korean Series 페이지"),
    ("MVP / 신인왕 / 골든글러브", "Wikipedia — 각 시상식 페이지"),
    ("", ""),
    ("알려진 한계", "2021-2025 선수 레벨 스탯은 미수집 (크롤러 ChromeDriver 재실행 필요)"),
    ("인코딩", "전 파일 UTF-8 BOM — Excel에서 바로 열림"),
    ("갱신 주기", "수동 — 시즌 종료 후 업데이트 권장"),
]
for i, (a, b) in enumerate(rows, start=3):
    if i == 3:
        for col, val in [(1, a), (2, b)]:
            c = s5.cell(row=i, column=col, value=val)
            c.font = HF; c.fill = HFILL; c.alignment = CEN
        s5.row_dimensions[i].height = 22
    else:
        s5.cell(row=i, column=1, value=a).font = Font(name="맑은 고딕", size=10, bold=True)
        s5.cell(row=i, column=2, value=b).font = BF
s5.column_dimensions["A"].width = 26
s5.column_dimensions["B"].width = 72

out = os.path.join(ROOT, "demo-test.xlsx")
wb.save(out)
print(f"OK {out}")
print(f"size: {os.path.getsize(out):,} bytes  sheets: {wb.sheetnames}")
